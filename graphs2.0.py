import openpyxl
from matplotlib import pyplot as plt

'''loads the file'''
def loadFile(name):
    workbook = openpyxl.load_workbook(str(name)+'.xlsx')
    return workbook

'''loads the specific sheet in the excel file'''
def loadSheet(sheetName):
    wb = loadFile(nameOfFile)
    sheet = wb[sheetName]
    return sheet

'''gets the titles from the excel file. It assumes
that the titles are in row one'''
def getTitles():
    titles = []
    sheet = loadSheet(sheetName)
    maxCol = sheet.max_column
    for i in range(2,maxCol):
        title = str(sheet.cell(row=1,column=i).value)
        titles.append(title)
    return titles    

'''determines the maximum amount of columns'''
def maxCol():
    sheet = loadSheet(sheetName)
    return sheet.max_column

'''determines the maximum amount of rows'''
def maxRow():
    sheet = loadSheet(sheetName)
    return sheet.max_row

'''loads the wavelengths in the specific sheet in the excel file'''
def loadWavelengths():
    wavelengths = []
    sheet = loadSheet(sheetName)
    maxRow = sheet.max_row
    for i in range(2,maxRow+1):
        wavelength = sheet.cell(row=i,column=1).value
        wavelengths.append(wavelength)
    return wavelengths

'''loads the absorbances in the specific sheet in the excel file
also subtracts the blank from each absorbance
program assumes blank is in the last column'''
def loadAbsorbances(columnNum):
    absorbances = []
    sheet = loadSheet(sheetName)
    maxRow = sheet.max_row
    maxCol = sheet.max_column
    for i in range(2,maxRow+1):
        absorbance = sheet.cell(row=i,column=columnNum).value
        blank = sheet.cell(row=i,column=maxCol).value
        absCorr = absorbance - blank
        absorbances.append(absCorr)
    return absorbances

'''adds all of the relavent columns to one graph'''
def allColumns():
    nameOfTitle = input('What do you want to name the title? ')
    for eachColumn in range(2,maxCol()):
        wavelengths = loadWavelengths()
        absorbances = loadAbsorbances(eachColumn)
        plt.plot(wavelengths,absorbances)
        plt.legend(getTitles())
        plt.title(nameOfTitle)
        plt.ylabel('Absorbance')
        plt.xlabel('Wavelength (nm)')
    plt.savefig(str(nameOfTitle)+'.png')
    
'''adds onely specific columns to one graph'''
def chooseColumn():
    whatColumn = input('What column do you want to choose? ')
    lengend = []
    while whatColumn != 'done':
        titleIndex = getTitles().index(whatColumn)
        lengend.append(getTitles()[titleIndex])
        wavelengths = loadWavelengths()
        absorbances = loadAbsorbances(titleIndex+2)
        plt.plot(wavelengths,absorbances)
        whatColumn = input('add another column or type done. ')
    titleName = input('What do you want the title to be? ')
    plt.title(titleName)
    plt.legend(lengend)
    plt.ylabel('Absorbance')
    plt.xlabel('Wavelength (nm)')
    ''''NOTE: getTitles() returns a list hence the [] for an index^^'''
    plt.savefig(str(titleName)+'.png')


def separateGraphs():
    for i in range(2,maxCol()):
        titleIndex = getTitles()[i-2]
        wavelengths = loadWavelengths()
        absorbances = loadAbsorbances(i)
        plt.plot(wavelengths,absorbances)
        plt.title(titleIndex)
        plt.ylabel('Absorbance')
        plt.xlabel('Wavelength (nm)')
        plt.savefig(str(titleIndex)+'.png')
        plt.close()
        
'''**********Main part of the program**********'''

nameOfFile = input('What is the name of the file? ')

'''lists all of the sheets and prints them'''
wb = loadFile(nameOfFile)
for i in range(len(wb.sheetnames)):
    print(wb.sheetnames[i])
    
sheetName = input('Which sheet do you want to use? ')
question = input('Separate or together? (s/t) ')

if question == 't':
    '''lists all of the titles in the sheet'''
    for i in range(len(getTitles())):
        print(getTitles()[i])

    question2 = input('all in this sheet? (y/n) ')
    if question2 == 'y':
        allColumns()
    else:
        print('Okay choose which one. ')
        chooseColumn()
else:
    separateGraphs()
    




